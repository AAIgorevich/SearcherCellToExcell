import os
from module.Comands import SCEComands, find_path_ini_file
import openpyxl
from tqdm import tqdm
from prettytable import PrettyTable
from time import sleep
import textwrap
import configparser


# Класс в котором присутсвуют инструменты для извлечение данных из конфиг файла
# или создания собственного конфига файла если последний отсутсвует
class ParserConfigToListOrCreateNew:

    # Хранилище переменных для внутреннего использования в классе
    def __init__(self) -> None:
        self.store_result = []  # Для
        self.files_and_path = {}
        self.config = configparser.ConfigParser()

    # чтение или создания конфиг файла если таковой отсутсвует
    def read_or_create_config(self) -> dict:
        if os.path.exists(file_config_ini):
            # Читаем конфиг файл
            self.config.read(file_config_ini)
            # Пробегаем по содержимому файла
            for section in self.config.sections():
                if self.config[section].values():
                    # Извлекаем путь
                    path = self.config[section]["path"].strip("''")
                    # Извлекаем наименование файлов
                    files = self.config[section]["files"]\
                        .strip("'").split(", ")
                    cleaned_files = \
                        [
                            (file.replace("'", ""))
                            .replace(",", "") for file in files
                        ]
                    # Создаем словарь куда помещаем данные
                    self.files_and_path.update({
                                section: {
                                    "path": path,
                                    "files": cleaned_files
                                    }
                            })
            return self.files_and_path
        else:  # Иначе создание конфигурационного файла
            print("config файл отсутсвует!")
            mark_str = ", ".join(
                [
                    f"'{element}'"
                    for element in self.search_excell_files_in_root()
                ]
                )
            new_config_file = open(file_config_ini, "w")
            if self.search_excell_files_in_root():
                new_config_file.write(textwrap.dedent("""
                # "ListGroups.GroupFile" Создался по причине того,
                # что в корневой папке программы присутсвуют файлы,
                # в которых можно осуществить поиск ячеек в Excell файлах.
                #  Если вы не желаете искать в этих файлах указанных в
                # "files", то просто удалите все начиная:
                # от "ListGroups.GroupFile", заканчивая "files"(включая).
                    """).strip()
                    )
                new_config_file.write(textwrap.dedent("""
                [ListGroups]
                [ListGroups.GroupFile]
                path = {}
                files = {}
                    """).format(
                        sce_workspace_dir,
                        mark_str))
            new_config_file.write(textwrap.dedent(
                r"""
                # Ниже представлен пример.
                # Раскоментируя его убрав "#",
                # Вы можете дополнить его или удалить
                # по собственному разумению.
                # [ListGroups.GroupFile1]
                # path = 'C:\Сюда_напишите_путь_к_файлу'
                # files = example_1.xlsx example_2.xlsx example_3.xlsx
                """).strip())
            new_config_file.close()
            print("Создан новый конфиг файл пожалуйста заполните его!")
            sleep(5)
            exit()

    #  Превращение словаря данных в тип list
    def parse_dict_to_list(self) -> list:
        # Получаем готовый словарь
        _file_path_dict: dict = self.read_or_create_config()
        # Проверяем пустой словарь или нет
        if _file_path_dict:
            self.store_result = [
                f"{group['path']}\\{file}"
                for group in _file_path_dict.values()
                for file in group['files']
            ]
        else:
            print("Конфиг файл не заполнен!")
            sleep(3)
            exit()
        _result = self.store_result
        return _result

    #  Поиск Excel файлов в корневой директории
    def search_excell_files_in_root(self) -> str:
        list_excell_files: list = []
        # Итерируемся по всем файлам в указанной папке
        for filename in os.listdir(sce_workspace_dir):
            if filename.endswith(".xlsx"):  # ! <= TODO *for old formats
                list_excell_files.append(f"{filename}")
        return list_excell_files


# Боевой класс.
# В нем описан поиск значений в ячейках excel.
class SCESearchInExcellFiles:

    # Хранилище переменных для внутреннего использования в классе
    def __init__(self) -> None:
        PCtL = ParserConfigToListOrCreateNew()
        self.list_path: list = PCtL.parse_dict_to_list()
        self.SCECommands = SCEComands()
        self.store_results: list = []
        self.user_input: str = ...
        self.str_found = \
            "\nНайдены совпадения в (.xlsx) файлах с вашем значением: "
        self.str_stroke = \
            "|===========================================================|"
        self.str_not_found = \
            "\nДанное значение не обнаруженно в (.xlsx) файлах."
        self.sv_tbl_str: str = ""
        self.clear_console_pointer: bool = False
        self.names_columns = [
                        "Имя файла",
                        "Название Листа",
                        "Координаты Ячейки"]

    # Поиск во всех листах excel файла
    def search_in_all_sheets(self, workbook, file_path):
        # Итерируемся по всем листам книги
        file_name = os.path.basename(file_path)
        for sheet_name in tqdm(
                workbook.sheetnames,
                # Вывод файлов которые были просмотренны
                desc=f"Просмотр файла {file_name}.",
                colour="#FFFF00",
                ascii=True, leave=False
                ):
            sheet = workbook[sheet_name]
            # Проверка на листы типа chart
            if sheet.title.startswith('Chart'):
                print(f"Пропускаем лист: {sheet.title}")
                continue
            self.search_in_all_cells(sheet, file_path, sheet_name)

    # Поиск во всех ячейках excel файла
    def search_in_all_cells(self, sheet, file_path, sheet_name):
        # Итерируемся по всем ячейкам в листе
        for row in sheet.iter_rows():
            for cell in row:
                # Проверяем
                # совпадает ли значение ячейки с искомыми
                cell_value = str(cell.value)
                if cell_value == self.user_input:
                    # Сохраняем путь к
                    # файлу, имени листа и адреса ячейки
                    self.store_results.append(
                        [os.path.basename(file_path),
                            sheet.title,
                            cell.coordinate])

    # Поиск в файле и загрузка книги excel файла
    def search_file_and_load_workbook(self):
        for file_path in self.list_path:
            # Загружаем книгу Excel
            workbook = openpyxl.load_workbook(
                file_path, read_only=True
                )
            self.search_in_all_sheets(workbook, file_path)

    # Боевая функция в ней происходят все процессы поиска
    def SCE_start_search_in_excel(self):
        try:
            self.SCECommands._first_init_command_help()
            while True:
                self.user_input = str(input("Ваше значение: "))
                command_result = \
                    self.SCECommands._call_comands(self.user_input)
                if command_result == "stop":
                    break
                if command_result == "continue":
                    self.sv_tbl_str = ""
                    continue
                if command_result == "save":
                    # сохраняем последний результат в файл
                    self.SCECommands._save_last_result_in_file(self.sv_tbl_str)
                    continue
                if isinstance(command_result, bool):
                    self.clear_console_pointer = command_result
                    continue
                if command_result is None:
                    pass
                func = {
                    True: (self.SCECommands._command_cleanup_console_output),
                    False: (lambda: None)
                }.get(self.clear_console_pointer)
                func()
                self.search_file_and_load_workbook()
                table = PrettyTable(self.names_columns)
                [table.add_row(row) for row in self.store_results]
                # Сохраняем в переменную чтобы можно было если нужно
                # сохранить результат в файл
                self.sv_tbl_str = table.get_string()
                if self.store_results:
                    print(self.str_found)
                    print(table)
                else:
                    print(self.str_not_found)
                print(self.str_stroke)
                self.store_results.clear()  # Очищаем список
        except KeyboardInterrupt:
            self.SCECommands._command_sce_stop()


if __name__ == "__main__":
    file_config_ini, sce_workspace_dir = find_path_ini_file()
    root = SCESearchInExcellFiles()
    root.SCE_start_search_in_excel()
