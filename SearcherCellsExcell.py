import os
import sys
import openpyxl
from tqdm import tqdm
from prettytable import PrettyTable
from time import sleep
import textwrap
import configparser


version = "1.1"
author = "AAIgorevich"
sce_workspace_dir = os.path.abspath(os.curdir)
file_config_ini = os.path.join(sce_workspace_dir, 'config.ini')


# Класс в котором сосредоточенны команды для программы
class SCEComands:
    # data text commands
    def __init__(self) -> None:
        self.SCE_greeting = textwrap.dedent(r"""
        ╔══════════════════════════════╗
        ║        ┌/\───/\┐             ║
        ║        │  SCE  │             ║
        ║        └──╗─╔──┘             ║
        ║ ╔═══──────╝─╚──────═══╗      ║
        ║ ║ ░ ╔──╗ ┌───┐ ╔──╗ ░ ║      ║
        ║ ╚═══╝  │ │>hi│ │  ╚═══╝      ║
        ║        │ └───┘ │             ║
        ║        └───────┘             ║
        ╚══════════════════════════════╝""").strip() 
        self.hint_help = textwrap.dedent("""
        Для того чтобы получить список команд,
        введите: 'searcher -help'.""").strip()
        self.stop_text = "Досвидания. Запускайте еще!"
        self.help_text = textwrap.dedent("""
        ╔══════════════════════════════════════════════╗
        ║                                              ║
        ║ Доступные команды:                           ║
        ║ ============================================ ║
        ║ searcher -hi  : Приветствие                  ║
        ║ searcher -stop: Выйти из программы           ║
        ║ searcher -info: Информация о программе       ║
        ║ searcher -help: Показать этот список команд  ║
        ║                                              ║
        ╚══════════════════════════════════════════════╝
        """).strip()
        self.hi_text = textwrap.dedent("""
        ╔════════════════════════════════════════════════════╗
        ║                                                    ║
        ║ Добро пожаловать в SearcherCellsExcell!            ║
        ║ ================================================== ║
        ║ Эта программа поможет вам быстро находить ячейки   ║
        ║ с определёнными значениями в ваших Excel файлах.   ║
        ║ Начните поиск и упростите свою работу с данными.   ║
        ║ Удачи!                                             ║
        ║                                                    ║
        ╚════════════════════════════════════════════════════╝
        """).strip()
        self.info_text = (textwrap.dedent("""
        ╔════════════════════════════════════════════════════╗
        ║                                                    ║
        ║ SearcherCellsExcell program information:           ║
        ║ ================================================== ║
        ║ SearcherCellsExcell or SCE version: {}            ║
        ║ author: {}                                ║
        ║ link GitHub author: https://github.com/{} ║
        ║                                                    ║
        ╚════════════════════════════════════════════════════╝
        """).strip()).format(version, author, author)

    # Остановка и выход из программы
    def command_sce_stop(self):
        print(self.stop_text)
        sleep(0.2)
        return sys.exit()

    # Вывод всех имеющихся команд на консоль (помощь)
    def command_sce_help(self):
        return print(self.help_text)

    # Приветствие на консоль
    def command_sce_hi(self):
        return print(self.hi_text)

    # Информациия о программе
    def command_sce_info(self):
        return print(self.info_text)

    # Подсказка для пользователей выводится единожды
    def first_init_command_help(self):
        print(self.SCE_greeting)
        sleep(0.2)
        return print(self.hint_help)

    # Вызов команд
    def call_comands(self, search_value) -> str | None:
        if search_value == "searcher -stop":
            self.command_sce_stop()
            return "stop"
        elif search_value == "searcher -help":
            self.command_sce_help()
            return "continue"
        elif search_value == "searcher -hi":
            self.command_sce_hi()
            return "continue"
        elif search_value == "searcher -info":
            self.command_sce_info()
            return "continue"
        # Если команда не распознана, возвращаем None, чтобы продолжить поиск
        return None


# Класс в котором присутсвуют инструменты для извлечение данных из конфиг файла
class ParserConfigToList:

    def __init__(self) -> None:
        self.store_result = []  # Для
        self.files_and_path = {}
        self.config = configparser.ConfigParser()

    def read_or_create_config(self) -> dict:
        if os.path.exists(file_config_ini):
            # Читаем конфиг файл
            self.config.read(file_config_ini)
            # Пробегаем по содержимому файла
            for section in self.config.sections():
                # Извлекаем путь
                path = self.config[section]["path"].strip("''")
                # Извлекаем наименование файлов
                files = self.config[section]["files"].split()
                # Создаем словарь куда помещаем данные
                self.files_and_path = {
                    "path": path,
                    "files": files
                }
        except Exception:
            print("config файл отсутсвует!")
        _result_dict = self.files_and_path
        return _result_dict

    def parse_dict_to_list(self):
        # Получаем готовый словарь
        _file_path_dict = self.try_read_config_and_write_in_dict()
        # Проверяем пустой словарь или нет
        if not _file_path_dict:
            self.store_result = [
                f"{group['path']}\\{file}"
                for group in _file_path_dict.values()
                for file in group['files']
            ]
        else:
            print("Конфиг файл не заполнен!")
        _result = self.store_result
        return _result


class SCESearchInExcellFiles:
    def __init__(self) -> None:
        PCtL = ParserConfigToList()
        self.list_path = PCtL.parse_dict_to_list()

    def read_path():
        pass

    def search_file(self):
        for filename in self.files_and_path:
            pass

    def loop_init(self):
        try:
            while True:
                input_search_value = str(input("Ваше значение: "))
                result = SCE.call_comands(input_search_value)
                if result == "stop":
                    break
                elif result == "continue":
                    continue
                elif result is None:
                    pass
                # ! <- Сюда ложим функции
        except KeyboardInterrupt:
            SCE.command_sce_stop()


SCE = SCEComands()

SCE.first_init_command_help()

# Укажите путь к папке с Excel-файлами
folder_path = os.path.abspath(os.curdir)


try:
    while True:
        # Укажите значение, которое нужно найти
        search_value = str(input("Ваше значение: "))

        # Получение результата комманды
        result = SCE.call_comands(search_value)

        # Остановка или
        # продолжение использование программы (зависит от "result")
        if result == "stop":
            break
        elif result == "continue":
            continue
        elif result is None:
            pass

        def find_cell_by_value(folder_path, search_value: str) -> list:
            store_results = []  # Список для хранения результатов поиска

            # Итерируемся по всем файлам в указанной папке
            for filename in tqdm(
                    os.listdir(folder_path),
                    desc="Статус работы: ",
                    colour="#00FF00", position=1, leave=False
                    ):
                if filename.endswith(".xlsx"):
                    file_path = os.path.join(folder_path, filename)
                    # Загружаем книгу Excel
                    workbook = openpyxl.load_workbook(
                        file_path, read_only=True
                        )
                    # Итерируемся по всем листам книги
                    for sheet_name in tqdm(
                            workbook.sheetnames,
                            # Вывод файлов которые были просмотренны
                            desc=f"Просмотр файла {filename}.",
                            colour="#FFFF00",
                            ascii=True, leave=False
                            ):
                        sheet = workbook[sheet_name]
                        # Итерируемся по всем ячейкам в листе
                        for row in sheet.iter_rows():
                            for cell in row:
                                # Проверяем
                                # совпадает ли значение ячейки с искомыми
                                cell_value = str(cell.value)
                                if cell_value == search_value:
                                    # Сохраняем путь к
                                    # файлу, имени листа и адреса ячейки
                                    store_results.append(
                                        [filename,
                                         sheet.title,
                                         cell.coordinate])

            # Возвращаем результаты поиска
            return store_results

        locations = find_cell_by_value(folder_path, search_value)
        table = PrettyTable(
            ["Имя файла", "Название Листа", "Координаты Ячейки"])
        for row in locations:
            table.add_row(row)
        if locations:
            print("\nНайдены совпадения в (.xlsx) файлах с вашем значением: ")
            print(table)
        else:
            print("\nДанное значение не обнаруженно в (.xlsx) файлах.")
        print("|===========================================================|")
except KeyboardInterrupt:
    SCE.command_sce_stop()
